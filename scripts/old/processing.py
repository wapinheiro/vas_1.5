import pandas as pd
import os
import glob           

def unify_source_files(source_folder, output_file): 
    # unify all source files into one csv file and save it in the same folder (source_data)
    # read and extract data from each file in the source_data folder
    # the source files have run_id information int he name of the file, it's the last three characters before the file extension.
    # Add a new column to the dataframe with the run_id extracted from the file name.
    # The output file will be saved in the same folder with the name 'unified_data.csv'

    print(f"Unifying source files from {source_folder} into {output_file}..."  )
    all_data = []
    
    for filename in os.listdir(source_folder):
        if filename.endswith('.csv'):
            file_path = os.path.join(source_folder, filename)
            run_id = filename[-7:-4]  # Extract run_id from the filename
            run_id = str(run_id).zfill(3)  # Always 3 digits, zero-padded
            df = pd.read_csv(file_path)
            df['run_id'] = run_id  # Add run_id column as zero-padded string
            all_data.append(df)
    
    unified_df = pd.concat(all_data, ignore_index=True)
    unified_df.to_csv(output_file, index=False) 

def create_cassette_table():
    """
    Create a cassette table from all run_parameters_template_*.xlsx files in the data/source_data folder.
    Each file is parsed for all cassettes (columns A/B, D/E, G/H, ...), extracting:
      - cassette_number (row 5)
      - cassette_code (row 6)
      - substrate_barcode (columns A, D, G, ...)
      - pallette_number (columns B, E, H, ...)
      - run (from filename, as number)
    The result is saved as data/processed/cassette_table.csv.
    """
    import os
    import glob
    import pandas as pd
    import openpyxl

    source_folder = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'source_data'))
    output_folder = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'processed'))
    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, 'cassette_table.csv')

    cassette_rows = []
    xlsx_files = glob.glob(os.path.join(source_folder, 'run_parameters_template_*.xlsx'))
    for xlsx_path in xlsx_files:
        run = int(os.path.basename(xlsx_path).split('_')[-1].split('.')[0])
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        # Find cassette columns: A/B, D/E, G/H, ...
        col_pairs = [(i, i+1) for i in range(1, ws.max_column, 3)]  # 1-based indexing
        for colA, colB in col_pairs:
            cassette_number = ws.cell(row=5, column=colA).value
            cassette_code = ws.cell(row=6, column=colA).value
            if not cassette_code:
                continue
            row = 7
            while True:
                substrate_barcode = ws.cell(row=row, column=colA).value
                pallette_number = ws.cell(row=row, column=colB).value
                if substrate_barcode is None and pallette_number is None:
                    break
                if substrate_barcode is not None and pallette_number is not None:
                    cassette_rows.append({
                        'cassette_number': cassette_number,
                        'cassette_code': cassette_code,
                        'substrate_barcode': str(substrate_barcode),
                        'pallette_number': pallette_number,
                        'run': run
                    })
                row += 1
    cassette_df = pd.DataFrame(cassette_rows)
    cassette_df.to_csv(output_file, index=False)
    print(f"Cassette table saved to {output_file}")

def create_pallette_table():
    pass

def create_view(source_file, output_file):
    import numpy as np
    # create the appropriate view for the chart and save it as a csv file in the 'views' folder
    # this will will contain the following columns:
    # Column A: 'TaskName2' (substrate barcode)
    # Column E: 'SpotNumber'
    # Column G: 'XOffset'
    # Column H: 'YOffset'
    # Column P: 'run_id' --> this should be a text column with the run_id information
    # Add: 'cassette_code' based on substrate_barcode and run_id from run_parameters_template_{run_id}.xlsx
    # This dataset should contain only unique rows based on the combination of 'TaskName2', 'SpotNumber', 'XOffset', and 'YOffset'.
    
    print(f"Creating view from {source_file} and saving to {output_file}...")       
    df = pd.read_csv(source_file)   
    # Format columns
    df['TaskName2'] = df['TaskName2'].astype(str)
    df['SpotNumber'] = pd.to_numeric(df['SpotNumber'], errors='coerce')
    df['XOffset'] = pd.to_numeric(df['XOffset'], errors='coerce')
    df['YOffset'] = pd.to_numeric(df['YOffset'], errors='coerce')
    df['run_id'] = df['run_id'].astype(str).str.zfill(3)

    # Track row count before cleaning
    before_rows = len(df)
    # Exclude rows where XOffset or YOffset is 999 (or close) or NaN
    df = df[(~np.isclose(df['XOffset'], 999)) &
            (~np.isclose(df['YOffset'], 999)) &
            (df['XOffset'].notnull()) &
            (df['YOffset'].notnull())]
    after_rows = len(df)
    print(f"Dropped {before_rows - after_rows} rows due to invalid XOffset/YOffset values.")

    view_df = df[['TaskName2', 'SpotNumber', 'XOffset', 'YOffset', 'run_id']].drop_duplicates()

    # Build a mapping: (run_id, substrate_barcode) -> cassette_code
    cassette_map = {}
    source_folder = os.path.dirname(source_file)
    xlsx_files = glob.glob(os.path.join(source_folder, 'run_parameters_template_*.xlsx'))
    for xlsx_path in xlsx_files:
        run_id = os.path.basename(xlsx_path)[-8:-5]
        xls = pd.ExcelFile(xlsx_path, engine='openpyxl')
        df_xls = pd.read_excel(xls, sheet_name=0, header=None)
        # Find cassette columns (headers in row 5, cassette codes in row 6)
        for col in range(df_xls.shape[1]):
            header = str(df_xls.iloc[4, col])
            cassette_code = str(df_xls.iloc[5, col])
            if header.startswith('Cassette') and cassette_code != 'nan':
                # Substrate barcodes start at row 6, cassette_code in this column
                for row in range(6, df_xls.shape[0]):
                    substrate_barcode = str(df_xls.iloc[row, col])
                    if substrate_barcode and substrate_barcode != 'nan':
                        cassette_map[(run_id, substrate_barcode)] = cassette_code

    # Map cassette_code to each row in the view
    def get_cassette(row):
        return cassette_map.get((row['run_id'], str(row['TaskName2'])), None)
    view_df['cassette_code'] = view_df.apply(get_cassette, axis=1)
    view_df['cassette_code'] = view_df['cassette_code'].astype(str)
    # Ensure run_id is saved as zero-padded string
    view_df['run_id'] = view_df['run_id'].astype(str).str.zfill(3)
    view_df.to_csv(output_file, index=False)  # Save the view as a CSV  file  

def create_view_case_02():
    pass

# Main execution
if __name__ == "__main__":
    
    source_folder = os.path.join('..', 'data', 'source_data')  # Corrected path to the source folder
    output_file = os.path.join(source_folder, 'unified_data.csv')  # Output file path for the unified data  
    
    # Unify source files
    # unify_source_files(source_folder, output_file)
    
    # Create view from the unified data
    # view_output_file = os.path.join('..', 'data', 'views', 'view_data.csv')  # Corrected path for the view output
    # create_view(output_file, view_output_file)  # Create the view from the unified data

    # Create cassette table
    create_cassette_table()
    
    print("Processing completed successfully.")


