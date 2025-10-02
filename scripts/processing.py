import os
import glob
import pandas as pd
import openpyxl


def unify_printing_files():
    """
    Unify all printing_raw_data_report_vas_*.csv files in data/source_data into one file.
    Adds a run_id column (int, from last 3 digits of filename) as the first column.
    Only common columns are kept. Skips empty files. Deduplicates rows.
    Output: data/processed/unified_printing.csv
    """

    source_folder = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'source_data'))
    output_folder = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'processed'))
    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, 'unified_printing.csv')

    all_dfs = []
    csv_files = glob.glob(os.path.join(source_folder, 'printing_raw_data_report_*.csv'))
    for file in csv_files:
        try:
            df = pd.read_csv(file)
        except Exception:
            continue
        if df.empty:
            continue
        run_id_str = os.path.basename(file).split('_')[-1].split('.')[0]
        # Handle both numeric and alphanumeric run_id values
        try:
            run_id = int(run_id_str[-5:])
        except ValueError:
            # If conversion to int fails, use the string as is
            run_id = run_id_str
        df.insert(0, 'run_id', run_id)
        all_dfs.append(df)
    if not all_dfs:
        print('No data to unify.')
        return
    # Only keep common columns
    common_cols = set(all_dfs[0].columns)
    for df in all_dfs[1:]:
        common_cols &= set(df.columns)
    common_cols = list(common_cols)
    all_dfs = [df[common_cols] for df in all_dfs]
    unified_df = pd.concat(all_dfs, ignore_index=True)
    unified_df = unified_df.drop_duplicates()
    unified_df.to_csv(output_file, index=False)
    print(f"Unified printing file saved to {output_file}")


def unify_run_parameters_files():
    """
    Unify all printing_raw_data_report_vas_*.csv files in data/source_data into one file.
    Adds a run_id column (int, from last 3 digits of filename) as the first column.
    Only common columns are kept. Skips empty files.
    Output: data/processed/unified_run_parameters.csv
    """

    source_folder = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'source_data'))
    output_folder = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'processed'))
    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, 'unified_run_parameters.csv')

    all_dfs = []
    csv_files = glob.glob(os.path.join(source_folder, 'printing_raw_data_report_*.csv'))
    for file in csv_files:
        try:
            df = pd.read_csv(file)
        except Exception:
            continue
        if df.empty:
            continue
        run_id_str = os.path.basename(file).split('_')[-1].split('.')[0]
        # Handle both numeric and alphanumeric run_id values
        try:
            run_id = int(run_id_str[-5:])
        except ValueError:
            # If conversion to int fails, use the string as is
            run_id = run_id_str
        df.insert(0, 'run_id', run_id)
        all_dfs.append(df)
    if not all_dfs:
        print('No data to unify.')
        return
    # Only keep common columns
    common_cols = set(all_dfs[0].columns)
    for df in all_dfs[1:]:
        common_cols &= set(df.columns)
    common_cols = list(common_cols)
    all_dfs = [df[common_cols] for df in all_dfs]
    unified_df = pd.concat(all_dfs, ignore_index=True)
    unified_df.to_csv(output_file, index=False)
    print(f"Unified run parameters file saved to {output_file}")


def create_cassette_table():
    """
    Create a cassette table from all run_parameters_template_*.xlsx files in the data/source_data folder.
    Each file is parsed for all cassettes (columns A/B, D/E, G/H, ...), extracting:
      - cassette_number (row 5)
      - cassette_code (row 6)
      - substrate_barcode (columns A, D, G, ...)
      - pallette_number (columns B, E, H, ...)
      - run (from filename, as number)
    The result is saved as data/processed/cassette_table.csv.
    Deduplicates rows based on (cassette_number, cassette_code, substrate_barcode, pallette_number, run).
    """

    source_folder = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'source_data'))
    output_folder = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'processed'))
    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, 'cassette_table.csv')

    cassette_rows = []
    xlsx_files = glob.glob(os.path.join(source_folder, 'run_parameters_template_*.xlsx'))
    for xlsx_path in xlsx_files:
        run_id_str = os.path.basename(xlsx_path).split('_')[-1].split('.')[0]
        # Handle both numeric and alphanumeric run_id values
        try:
            run = int(run_id_str)
        except ValueError:
            # If conversion to int fails, use the string as is
            run = run_id_str
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        # Find cassette columns: A/B, D/E, G/H, ...
        col_pairs = [(i, i+1) for i in range(1, ws.max_column, 3)]  # 1-based indexing
        for colA, colB in col_pairs:
            cassette_number = ws.cell(row=5, column=colA).value
            cassette_code = ws.cell(row=6, column=colA).value
            if not cassette_code:
                continue
            row = 7
            while True:
                substrate_barcode = ws.cell(row=row, column=colA).value
                pallette_number = ws.cell(row=row, column=colB).value
                if substrate_barcode is None and pallette_number is None:
                    break
                if substrate_barcode is not None and pallette_number is not None:
                    cassette_rows.append({
                        'cassette_number': cassette_number,
                        'cassette_code': cassette_code,
                        'substrate_barcode': str(substrate_barcode),
                        'pallette_number': pallette_number,
                        'run': run
                    })
                row += 1
    cassette_df = pd.DataFrame(cassette_rows)
    # Deduplicate rows
    cassette_df = cassette_df.drop_duplicates()
    cassette_df.to_csv(output_file, index=False)
    print(f"Cassette table saved to {output_file}")


def create_pallette_table():
    """
    Create a pallette table from cassette_table.csv, extracting only (run, pallette_number, substrate_barcode),
    deduplicating rows, and saving as data/processed/pallette_table.csv.
    """

    input_file = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'processed', 'cassette_table.csv'))
    output_folder = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'processed'))
    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, 'pallette_table.csv')

    if not os.path.exists(input_file):
        print(f"Cassette table not found: {input_file}")
        return

    df = pd.read_csv(input_file, dtype=str)  # Use dtype=str to avoid type issues
    cols = ['run', 'pallette_number', 'substrate_barcode']
    missing = [c for c in cols if c not in df.columns]
    if missing:
        print(f"Missing columns in cassette_table.csv: {missing}")
        return
    pallette_df = df[cols].drop_duplicates().reset_index(drop=True)
    pallette_df.to_csv(output_file, index=False)
    print(f"Pallette table saved to {output_file}")


def ensure_cassette_table_exists():
    """
    Checks if data/processed/cassette_table.csv exists. If not, runs create_cassette_table().
    """
    output_file = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'processed', 'cassette_table.csv'))
    if not os.path.exists(output_file):
        print(f"Cassette table not found at {output_file}. Creating...")
        create_cassette_table()
    else:
        print(f"Cassette table found at {output_file}.")


def create_view_case_01():
    """
    Creates a view for case 1 from unified_printing.csv.
    - Loads data/processed/unified_printing.csv
    - Selects columns: run_id, SpotNumber, XOffset, YOffset
    - Deduplicates rows
    - Saves as data/views/view_case_01.csv
    """

    printing_file = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'processed', 'unified_printing.csv'))
    output_folder = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'views'))
    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, 'view_case_01.csv')

    # Load data
    cols = ['run_id', 'SpotNumber', 'XOffset', 'YOffset']
    df = pd.read_csv(printing_file, dtype=str)
    df = df[[col for col in cols if col in df.columns]]
    df = df.drop_duplicates()
    df.to_csv(output_file, index=False)
    print(f"View for case 1 saved to {output_file}")


def create_view_case_02():
    """
    Creates a view for case 2 by joining unified_printing.csv and pallette_table.csv.
    - Loads data/processed/unified_printing.csv and data/processed/pallette_table.csv
    - Selects columns: run_id, SpotNumber, XOffset, YOffset, TaskName2 from unified_printing.csv
    - Selects pallette_number, substrate_barcode from pallette_table.csv
    - Left joins on unified_printing.TaskName2 == pallette_table.substrate_barcode
    - Keeps all matches (if multiple)
    - Outputs columns: run_id, SpotNumber, XOffset, YOffset, pallette_number
    - Saves as data/views/view_case_02.csv
    """

    printing_file = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'processed', 'unified_printing.csv'))
    pallette_file = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'processed', 'pallette_table.csv'))
    output_folder = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'views'))
    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, 'view_case_02.csv')

    # Load data
    printing_cols = ['run_id', 'SpotNumber', 'XOffset', 'YOffset', 'TaskName2']
    pallette_cols = ['pallette_number', 'substrate_barcode']
    printing_df = pd.read_csv(printing_file, dtype=str)
    pallette_df = pd.read_csv(pallette_file, dtype=str)

    # Filter columns
    printing_df = printing_df[[col for col in printing_cols if col in printing_df.columns]]
    pallette_df = pallette_df[[col for col in pallette_cols if col in pallette_df.columns]]

    # Left join
    merged = printing_df.merge(pallette_df, how='left', left_on='TaskName2', right_on='substrate_barcode')

    # Select output columns
    out_cols = ['run_id', 'SpotNumber', 'XOffset', 'YOffset', 'pallette_number']
    merged[out_cols].to_csv(output_file, index=False)
    print(f"View for case 2 saved to {output_file}")

def create_view_case_03():
    """
    Creates a view for case 3 by joining unified_printing.csv and pallette_table.csv.
    - Loads data/processed/unified_printing.csv and data/processed/pallette_table.csv
    - Selects columns: run_id, RowNumber, SpotNumber, XOffset, YOffset, TaskName2 from unified_printing.csv
    - Selects pallette_number, substrate_barcode from pallette_table.csv
    - Left joins on unified_printing.TaskName2 == pallette_table.substrate_barcode
    - Keeps all matches (if multiple)
    - Outputs columns: run_id, SpotNumber, XOffset, YOffset, pallette_number
    - Saves as data/views/view_case_02.csv
    """

    printing_file = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'processed', 'unified_printing.csv'))
    pallette_file = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'processed', 'pallette_table.csv'))
    cassette_file = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'processed', 'cassette_table.csv'))
    output_folder = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'views'))
    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, 'view_case_03.csv')

    # Load data
    printing_cols = ['run_id', 'RowNumber', 'SpotNumber', 'XOffset', 'YOffset', 'TaskName2']
    pallette_cols = ['pallette_number', 'substrate_barcode']
    cassette_cols = ['cassette_number', 'substrate_barcode', 'run']

    printing_df = pd.read_csv(printing_file, dtype=str)
    pallette_df = pd.read_csv(pallette_file, dtype=str)
    cassette_df = pd.read_csv(cassette_file, dtype=str)

    # Filter columns
    printing_df = printing_df[[col for col in printing_cols if col in printing_df.columns]]
    pallette_df = pallette_df[[col for col in pallette_cols if col in pallette_df.columns]]
    cassette_df = cassette_df[[col for col in cassette_cols if col in cassette_df.columns]]


    # Left join
    merged = printing_df.merge(pallette_df, how='left', left_on='TaskName2', right_on='substrate_barcode')
    merged = merged.merge(cassette_df, how='left', left_on=['run_id', 'TaskName2'], right_on=['run', 'substrate_barcode'])

    # Select output columns
    out_cols = ['run_id', 'RowNumber', 'SpotNumber', 'XOffset', 'YOffset', 'cassette_number', 'pallette_number']
    merged[out_cols].to_csv(output_file, index=False)
    print(f"View for case 3 saved to {output_file}")


# Main execution
if __name__ == "__main__":  
    unify_printing_files()    
    unify_run_parameters_files()
    create_cassette_table()
    create_pallette_table()
    create_view_case_01()
    create_view_case_02()
    create_view_case_03()
